# Importa as bibliotecas necessárias
import numpy as np
import pandas as pd
import os
import cv2
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# Dicionário para mapear rótulos de reconhecimento facial a nomes de alunos
names = {0: 'n', 1: 'm', 2: 'v', 3: 'l'}

def verify_and_load_workbook(filename):
    """
    Verifica se o arquivo existe e se é um arquivo Excel (.xlsx).
    
    Parâmetros:
    - filename: Nome do arquivo a ser verificado e carregado
    
    Retorna:
    - workbook: Objeto Workbook do openpyxl se o carregamento for bem-sucedido
    """
    if not os.path.isfile(filename):
        raise FileNotFoundError(f"Arquivo não encontrado: {filename}")
    if not filename.endswith('.xlsx'):
        raise ValueError(f"Arquivo não é um arquivo Excel (.xlsx): {filename}")
    
    try:
        workbook = load_workbook(filename=filename)
        print(f"Arquivo {filename} carregado com sucesso.")
        return workbook
    except Exception as e:
        raise ValueError(f"Erro ao carregar o arquivo {filename}: {e}")


class detect_image:
    def __init__(self, recordSheet, attendenceSheet):
        """
        Inicializa a classe de detecção de imagem com planilhas de registro e presença.
        
        Parâmetros:
        - recordSheet: Nome do arquivo Excel de registro
        - attendenceSheet: Nome do arquivo Excel de presença
        """
        # Inicializa os parâmetros da classe
        self.recordSheet = recordSheet
        self.attendenceSheet = attendenceSheet
        
        # Carrega o classificador de faces Haar Cascade
        self.cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
        
        # Inicializa o reconhecedor facial LBPH e carrega um modelo treinado
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.recognizer.read('trained.yml')
        
        # Lista para armazenar alunos já marcados como presentes
        self.students_marked = []
        
        # Carrega as planilhas de registro e de presença
        self.recordWorkbook = verify_and_load_workbook(filename=recordSheet)
        self.attendenceWorkbook = verify_and_load_workbook(filename=attendenceSheet)

    def identify(self):
        """
        Identifica um aluno usando reconhecimento facial em tempo real.
        
        Retorna:
        - Nome do aluno identificado
        """
        # Inicializa a captura de vídeo
        cap = cv2.VideoCapture(0)
        f = 0  # Contador de quadros
        pred_name = "None"  # Nome predito do aluno
        
        while True:
            # Captura um quadro de vídeo
            _, frame = cap.read()
            
            # Converte o quadro para escala de cinza
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            # Equaliza o histograma da imagem em escala de cinza
            equalize = cv2.equalizeHist(gray_frame)
            
            # Aplica filtro de mediana para reduzir ruído
            image = cv2.medianBlur(equalize, 3)
            
            # Detecta faces na imagem
            face = self.cascade.detectMultiScale(image, 1.1, 5)
            
            for x, y, w, h in face:
                # Desenha um retângulo em torno da face detectada
                cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 255, 0), 4)
                
                # Prediz o rótulo e a confiança para a face detectada
                label, conf = self.recognizer.predict(gray_frame[y:y+h, x:x+w])
                
                # Obtém o nome correspondente ao rótulo predito
                p_name = names.get(label)
                pred_name = p_name
                
                # Exibe o quadro com a face reconhecida
                cv2.imshow("Recognizing Face", frame)
                
                # Verifica se a tecla 'q' foi pressionada para interromper a captura de vídeo
                flag = cv2.waitKey(1) & 0xFF
                f += 1
                
                # Se mais de 15 quadros foram processados, libera a captura de vídeo e retorna o nome predito
                if f > 15:
                    cap.release()
                    return p_name

    def writeEntry(self, p_name, branch, year):
        """
        Escreve a entrada de presença na planilha de presença.
        
        Parâmetros:
        - p_name: Nome do aluno
        - branch: Ramo de estudo do aluno
        - year: Ano do aluno
        """
        # Escreve a entrada de presença na planilha
        sheet = self.attendenceWorkbook.active
        sheet.insert_rows(idx=2, amount=1)
        sheet["A2"] = p_name
        sheet["B2"] = year
        sheet["C2"] = branch
        
        # Obtém o horário atual e escreve na planilha
        time = datetime.now()
        curr_time = time.strftime("%H:%M:%S")
        sheet["D2"] = curr_time
        
        # Salva a planilha de presença
        self.attendenceWorkbook.save(self.attendenceSheet)

    def markEntry(self, p_name):
        """
        Marca a entrada de um aluno na planilha de presença.
        
        Parâmetros:
        - p_name: Nome do aluno
        """
        # Marca a entrada de um aluno na planilha de presença
        if p_name in self.students_marked:
            print("Entrada já marcada!")
        else:
            self.students_marked.append(p_name)
            sheet = self.recordWorkbook.active
            
            # Procura o nome do aluno na planilha de registros e obtém seus dados
            for values in sheet.iter_rows(min_col=1, max_col=3, values_only=True):
                rec_name = values[0]
                if rec_name.lower() == p_name:
                    branch = values[1]
                    year = values[2]
                    self.writeEntry(p_name, branch, year)

    def close_entry(self, p_name):
        """
        Fecha a entrada de presença de um aluno na planilha de presença.
        
        Parâmetros:
        - p_name: Nome do aluno
        """
        # Fecha a entrada de presença de um aluno na planilha de presença
        if not p_name in self.students_marked:
            print("Entrada não marcada!")
        else:
            sheet = self.attendenceWorkbook.active
            index = 0
            
            # Procura o nome do aluno na planilha de presença
            for values in sheet.iter_rows(values_only=True):
                rec_name = values[0]
                index += 1
                if rec_name.lower() == p_name:
                    # Obtém o horário atual e escreve na planilha
                    time = datetime.now()
                    curr_time = time.strftime("%H:%M:%S")
                    sheet["E" + str(index)] = curr_time
                    print("Entrada fechada!")
                    
                    # Salva a planilha de presença e remove o aluno da lista de alunos marcados
                    self.attendenceWorkbook.save(self.attendenceSheet)
                    self.students_marked.remove(p_name)
                    break
